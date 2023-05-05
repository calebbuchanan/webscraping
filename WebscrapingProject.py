from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font, PatternFill
import keys
from twilio.rest import Client


url = 'https://finance.yahoo.com/u/yahoo-finance/watchlists/crypto-top-market-cap/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}


req = Request(url, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')

wb = xl.Workbook()
ws = wb.active
ws.title = 'CryptoReport'

ws['A1'] = 'Ticker Symbol'
ws['A1'].font = Font(name='Times New Roman', size=18, bold=True, italic=False)

ws['B1'] = 'Name'
ws['B1'].font = Font(name='Times New Roman', size=18, bold=True, italic=False)

ws['C1'] = 'Current Price'
ws['C1'].font = Font(name='Times New Roman', size=18, bold=True, italic=False)

ws['D1'] = 'Percect Change over 24 Hours'
ws['D1'].font = Font(name='Times New Roman', size=18, bold=True, italic=False)

ws['E1'] = 'Original Price'
ws['E1'].font = Font(name='Times New Roman', size=18, bold=True, italic=False)


crypto_data = soup.findAll('tr')

counter = 1
for x in range(1, 6):
    td = crypto_data[3+counter].findAll('td')

    symbol = td[0].text
    name = td[1].text
    current_price = float(td[2].text)
    percent_change24hrs = float(td[4].text.strip("%"))/100
    original_price = current_price - (current_price * (percent_change24hrs))

    ws['A' + str(x+1)] = symbol
    ws['B' + str(x+1)] = name
    ws['C' + str(x+1)] = '$' + format(current_price, ',.2f')
    ws['D' + str(x+1)] = format(percent_change24hrs, '.2%')
    ws['E' + str(x+1)] = '$' + format(original_price, ',.2f')

    counter += 1


ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 25

ws['A1'].fill = PatternFill(patternType='solid', fgColor='90EE90')
ws['B1'].fill = PatternFill(patternType='solid', fgColor='90EE90')
ws['C1'].fill = PatternFill(patternType='solid', fgColor='90EE90')
ws['D1'].fill = PatternFill(patternType='solid', fgColor='90EE90')
ws['E1'].fill = PatternFill(patternType='solid', fgColor='90EE90')



wb.save('CryptocurrencyReport.xlsx')

#Twilio Code

client = Client(keys.accountSID, keys.authToken)

TwilioNumber = '19727377197'
mycellphone = '+19723960213'

checkcurrency = print('Check Cryptocurrencies')
changeinprice = current_price - original_price

if (symbol == 'BTC-USD' or symbol == 'ETH-USD') and (changeinprice > 5 or changeinprice < -5):
    textmsg = client.messages.create(to=mycellphone, from_=TwilioNumber, body=checkcurrency)
